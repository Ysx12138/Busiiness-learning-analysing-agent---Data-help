"""数据分析工具集 —— 数据概览和交付物生成。"""

import csv
import html
import os
import re
import statistics
from pathlib import Path


def read_csv_summary(path: str, repo_root: str) -> str:
    """读取 CSV 文件并返回基础概览：行数、列数、列名。"""
    from datahelp.tools import resolve_path
    path = resolve_path(path, repo_root)
    p = Path(path)
    if not p.exists():
        return f"错误: 文件不存在 {p}"
    if not p.is_file():
        return f"错误: {p} 不是文件"

    try:
        with p.open(newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
    except Exception as e:
        return f"错误: 无法读取 CSV 文件 - {e}"

    if not rows:
        return f"CSV 文件 {p} 是空的"

    headers = rows[0]
    col_count = len(headers)
    data_rows = len(rows) - 1

    lines = [
        f"文件: {p}",
        f"总行数: {data_rows}",
        f"总列数: {col_count}",
        f"列名: {', '.join(headers)}",
    ]
    return "\n".join(lines)


def read_evidence_file(path: str) -> str:
    """读取 UTF-8 证据文件，返回内容（最多 30000 字符）。不存在或无法读取返回空字符串。"""
    p = Path(path)
    if not p.exists() or not p.is_file():
        return ""
    try:
        return p.read_text(encoding="utf-8", errors="replace")[:30000]
    except (OSError, ValueError):
        return ""


# ── 交付物生成 ────────────────────────────────────────

def _load_csv_data(path: str, repo_root: str) -> tuple[list[str], list[dict]]:
    """加载 CSV 文件，返回 (列名列表, 数据行列表)。"""
    from datahelp.tools import resolve_path
    path = resolve_path(path, repo_root)
    p = Path(path)
    with p.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    return (reader.fieldnames or []), rows


def _is_numeric(v: str) -> bool:
    """检查字符串是否为数值。"""
    try:
        float(v)
        return True
    except ValueError:
        return False


def _infer_numeric_cols(headers: list[str], rows: list[dict]) -> list[str]:
    """推断数值列。"""
    numeric = []
    for col in headers:
        vals = [r.get(col, "").strip() for r in rows if r.get(col, "").strip()]
        if not vals:
            continue
        num_count = sum(1 for v in vals if _is_numeric(v))
        if num_count / len(vals) > 0.9:
            numeric.append(col)
    return numeric


def _get_categorical_cols(headers: list[str], rows: list[dict], numeric_cols: list[str]) -> list[str]:
    """推断类别列（非数值 + 唯一值数量适中）。"""
    numeric_set = set(numeric_cols)
    cat_cols = []
    n = len(rows)
    max_unique = max(5, min(50, n // 3))  # 至少 5，最多 50，且不超过 1/3 行数
    for col in headers:
        if col in numeric_set:
            continue
        vals = [r.get(col, "").strip() for r in rows if r.get(col, "").strip()]
        unique = len(set(vals))
        if 2 <= unique <= max_unique:
            cat_cols.append(col)
    return cat_cols


def _extract_section(text: str, section_keywords: list[str], max_chars: int = 800) -> str:
    """从分析文本中提取包含指定关键词的段落。"""
    if not text:
        return ""
    lines = text.split("\n")
    result = []
    capture = False
    for line in lines:
        if any(kw in line for kw in section_keywords):
            capture = True
        if capture:
            result.append(line)
            if len("".join(result)) > max_chars:
                break
    return "\n".join(result)


def _parse_report_sections(text: str) -> dict[str, str]:
    """轻量 Markdown 报告章节解析器。

    将分析报告按 ## / # 标题拆分为 {章节名: 内容} 字典。
    不引入新依赖，纯文本解析。

    返回:
        {章节名: 章节内容} 的字典。首个标题之前的内容以 "_preamble" 为键。
    """
    if not text:
        return {}
    sections: dict[str, str] = {}
    current_section = "_preamble"
    current_lines: list[str] = []

    for line in text.split("\n"):
        if line.startswith("## ") or line.startswith("# "):
            if current_lines:
                sections[current_section] = "\n".join(current_lines).strip()
            current_section = line.lstrip("# ").strip()
            current_lines = []
        else:
            current_lines.append(line)

    if current_lines:
        sections[current_section] = "\n".join(current_lines).strip()

    return sections


def _get_section_aliases() -> dict[str, list[str]]:
    """返回业务章节的可能 Markdown 标题变体。"""
    return {
        "执行摘要": ["执行摘要", "Executive Summary", "概述", "概要"],
        "数据概览": ["数据概览", "数据概述", "数据集概览", "Data Overview"],
        "核心发现": ["核心发现", "关键发现", "主要发现", "Key Findings"],
        "业务建议": ["业务建议", "行动建议", "建议", "Recommendations", "Action Items"],
        "数据质量": ["数据质量检查", "数据质量", "Data Quality"],
        "基础指标": ["基础指标分析", "描述性统计", "Basic Metrics"],
    }


def _extract_top_findings(text: str, max_count: int = 3) -> list[str]:
    """从 Markdown 文本中提取前 N 个关键发现项。

    匹配编号列表 (1. / 2.) 或 破折号列表 (-)。
    若找不到编号或短横线列表项，回退使用该章节首 max_count 行非空正文。
    """
    findings: list[str] = []
    for line in text.split("\n"):
        stripped = line.strip()
        # 匹配 "1. **xxx**" 或 "- xxx" 或 "**xxx**" 格式
        if re.match(r'^\d+\.\s', stripped):
            findings.append(stripped)
        elif stripped.startswith("- ") and len(stripped) > 3:
            findings.append(stripped[2:].strip())
        elif stripped.startswith("**") and stripped.endswith("**") and len(stripped) > 6:
            findings.append(stripped.strip("*"))
        if len(findings) >= max_count:
            break

    # 回退: 若找不到格式化列表项，取首 max_count 行非空正文
    if not findings:
        for line in text.split("\n"):
            stripped = line.strip()
            if stripped and not stripped.startswith("#"):
                findings.append(stripped)
                if len(findings) >= max_count:
                    break

    return findings


def _clean_markdown(text: str) -> str:
    """去除 Markdown 标记：#、**、-、>，替换 ⚠ 为“注意：”，移除 U+FE0F。"""
    if not text:
        return ""
    text = text.replace("⚠", "注意：").replace("️", "")
    lines = text.split("\n")
    cleaned = []
    for line in lines:
        line = re.sub(r'^#{1,6}\s+', '', line)
        line = line.replace('**', '')
        line = re.sub(r'^[-*]\s+', '', line)
        line = re.sub(r'^>\s+', '', line)
        if re.match(r'^\|\s*[-:| ]+\s*\|$', line):
            continue
        line = line.strip()
        if line:
            cleaned.append(line)
    return "\n".join(cleaned)


def _extract_structured_summary(analysis_text: str) -> dict:
    """从分析文本中提取结构化摘要。

    返回包含以下键的字典:
        summary: 执行摘要文本（优先"执行摘要"章节，否则首个非标题段落）
        findings: 核心发现列表（最多 3 条，仅从"核心发现/关键发现"章节提取）
        recommendations: 行动建议文本（从"业务建议/行动建议"章节提取）

    所有文本均已去除 Markdown 标记（#、**、-），保留序号可读文本。
    不要把整个 ## 标题塞入摘要。
    """
    if not analysis_text:
        return {"summary": "", "findings": [], "recommendations": ""}

    sections = _parse_report_sections(analysis_text)

    # ── 1. 执行摘要 ──
    summary = ""
    for key in ("执行摘要", "Executive Summary", "概述", "概要"):
        if sections.get(key, "").strip():
            summary = sections[key]
            break
    if not summary:
        # 没有专门的执行摘要章节，取首个非 Markdown 标题的段落
        paragraphs = [p.strip() for p in analysis_text.split("\n\n") if p.strip()]
        for p in paragraphs:
            if not re.match(r'^#+\s', p):
                summary = p
                break
    summary = _clean_markdown(summary)
    if len(summary) > 800:
        summary = summary[:797] + "..."

    # ── 2. 核心发现（仅从 核心发现/关键发现 章节提取，不可从整篇乱取） ──
    findings = []
    findings_text = ""
    for key in ("核心发现", "关键发现", "主要发现", "Key Findings"):
        if sections.get(key, "").strip():
            findings_text = sections[key]
            break
    if findings_text:
        raw = _extract_top_findings(findings_text, 3)
        findings = [_clean_markdown(f) for f in raw[:3]]
        if not findings:
            # 回退：遍历行，清理非空行，跳过纯标题，取前三条
            for line in findings_text.splitlines():
                cleaned = _clean_markdown(line).strip()
                if cleaned and not line.strip().startswith("#"):
                    findings.append(cleaned)
                    if len(findings) >= 3:
                        break

    # ── 3. 行动建议获取 ──
    recommendations = ""
    for key in ("业务建议", "行动建议", "建议", "Recommendations", "Action Items"):
        if sections.get(key, "").strip():
            recommendations = sections[key]
            break
    if not recommendations:
        fallback = _extract_section(analysis_text, ["建议", "推荐", "行动", "Recommend", "Action"], 1000)
        if fallback:
            recommendations = fallback
    recommendations = _clean_markdown(recommendations)

    return {
        "summary": summary,
        "findings": findings,
        "recommendations": recommendations,
    }


def _add_dimension_sheet(wb, headers, rows, numeric_cols, header_font, header_fill, thin_border):
    """添加维度分析工作表：自动检测类别列，每个类别列一张分组统计表。"""
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    cat_cols = _get_categorical_cols(headers, rows, numeric_cols)
    if not cat_cols or not numeric_cols:
        return

    ws_dim = wb.create_sheet("维度分析")

    # 对每个类别列，计算各组的数值列均值
    current_row = 1
    for cat_col in cat_cols[:3]:  # 最多 3 个维度，避免过多 sheet
        # 分组聚合
        groups: dict[str, list[dict]] = {}
        for r in rows:
            val = r.get(cat_col, "").strip()
            if val:
                groups.setdefault(val, []).append(r)

        # 计算每组均值
        dim_data: list[list] = [["分组"] + [f"{num}_均值" for num in numeric_cols[:5]]]
        for group_name, group_rows in sorted(groups.items())[:10]:
            row_data = [group_name]
            for num_col in numeric_cols[:5]:
                vals = []
                for r in group_rows:
                    v = r.get(num_col, "").strip()
                    if v and _is_numeric(v):
                        vals.append(float(v))
                row_data.append(round(statistics.mean(vals), 2) if len(vals) >= 1 else "")
            dim_data.append(row_data)

        # 标题
        cell = ws_dim.cell(row=current_row, column=1, value=f"维度: {cat_col}")
        cell.font = openpyxl.styles.Font(bold=True, size=12, color="4472C4")
        current_row += 1

        # 写入表格
        for r, row_data in enumerate(dim_data, current_row):
            for c, val in enumerate(row_data, 1):
                cell = ws_dim.cell(row=r, column=c, value=val)
                if r == current_row:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.border = thin_border

        current_row += len(dim_data) + 1

    for c in range(1, 7):
        ws_dim.column_dimensions[get_column_letter(c)].width = 18


def _add_cross_analysis_sheet(wb, headers, rows, numeric_cols, header_font, header_fill, thin_border):
    """添加交叉分析工作表：2D 透视表。"""
    import openpyxl
    from openpyxl.utils import get_column_letter

    cat_cols = _get_categorical_cols(headers, rows, numeric_cols)
    if len(cat_cols) < 2 or not numeric_cols:
        return

    ws_cross = wb.create_sheet("交叉分析")

    dim1, dim2 = cat_cols[0], cat_cols[1]
    target = numeric_cols[0]

    # 收集行列标签
    labels1 = sorted(set(r.get(dim1, "").strip() for r in rows if r.get(dim1, "").strip()))
    labels2 = sorted(set(r.get(dim2, "").strip() for r in rows if r.get(dim2, "").strip()))

    # 限制大小避免 Excel 崩溃
    labels1 = labels1[:10]
    labels2 = labels2[:10]

    # 计算交叉均值
    pivot_data: dict[tuple[str, str], list[float]] = {}
    for r in rows:
        v1 = r.get(dim1, "").strip()
        v2 = r.get(dim2, "").strip()
        val = r.get(target, "").strip()
        if v1 in labels1 and v2 in labels2 and val and _is_numeric(val):
            pivot_data.setdefault((v1, v2), []).append(float(val))

    # 标题
    cell = ws_cross.cell(row=1, column=1, value=f"交叉分析: {dim1} × {dim2} (均值 {target})")
    cell.font = openpyxl.styles.Font(bold=True, size=12, color="4472C4")

    # 表头
    ws_cross.cell(row=2, column=1, value=f"{dim1} \\ {dim2}")
    ws_cross.cell(row=2, column=1).font = header_font
    ws_cross.cell(row=2, column=1).fill = header_fill
    for j, l2 in enumerate(labels2, 2):
        cell = ws_cross.cell(row=2, column=j, value=l2)
        cell.font = header_font
        cell.fill = header_fill

    # 数据体
    for i, l1 in enumerate(labels1, 3):
        ws_cross.cell(row=i, column=1, value=l1)
        ws_cross.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True)
        for j, l2 in enumerate(labels2, 2):
            key = (l1, l2)
            if key in pivot_data and pivot_data[key]:
                vals = pivot_data[key]
                val = round(statistics.mean(vals), 2)
                ws_cross.cell(row=i, column=j, value=val)

    for c in range(1, len(labels2) + 2):
        ws_cross.column_dimensions[get_column_letter(c)].width = 18


def _add_thinking_models_sheet(wb, analysis_text, header_font, header_fill, thin_border):
    """添加思维模型工作表：5 种思维模型的教学内容。"""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws_tm = wb.create_sheet("思维模型")

    thinking_models = [
        ("1. 分解", "把总量拆解为有意义的组成部分（按时间、品类、区域、渠道等）",
         "拆解后分析各组成部分的贡献和差异"),
        ("2. 分层差异", "不盲目相信平均值，比较子群间的差异（包括倍数/比率）",
         "识别哪些子群表现明显不同，理解差异来源"),
        ("3. 代理推断", "当关键概念没有直接记录时，用可观测的代理信号推断",
         "例如用 AOV 推断消费力，用复购频率推断忠诚度"),
        ("4. 约束 vs 偏好", "判断差异来自用户不能选择（约束）还是不愿选择（偏好）",
         "约束暗示产品/流程改进，偏好暗示营销/定位调整"),
        ("5. 杠杆点", "聚焦高份额 + 高改进空间的细分领域",
         "杠杆价值 ≈ 受影响用户占比 × 潜在改进幅度"),
    ]

    # 标题
    ws_tm.cell(row=1, column=1, value="思维模型教学").font = Font(bold=True, size=14)
    ws_tm.cell(row=1, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_tm.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws_tm.merge_cells("A1:D1")

    # 表头
    headers_tm = ["思维模型", "核心思想", "应用场景", "数据证据"]
    for c, h in enumerate(headers_tm, 1):
        cell = ws_tm.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # 数据行
    for i, (name, desc, example) in enumerate(thinking_models, 3):
        ws_tm.cell(row=i, column=1, value=name).border = thin_border
        ws_tm.cell(row=i, column=2, value=desc).border = thin_border

        # 从 analysis_text 中提取相关证据
        evidence = ""
        if analysis_text:
            kw = name.split(".", 1)[1].strip() if "." in name else name
            evidence = _extract_section(analysis_text, [kw, desc[:8]], 300)

        ws_tm.cell(row=i, column=3, value=example).border = thin_border
        ws_tm.cell(row=i, column=4, value=evidence if evidence else "本次分析未应用").border = thin_border

    ws_tm.column_dimensions["A"].width = 18
    ws_tm.column_dimensions["B"].width = 40
    ws_tm.column_dimensions["C"].width = 40
    ws_tm.column_dimensions["D"].width = 50


def _add_self_check_sheet(wb, analysis_text, header_font, header_fill, thin_border):
    """添加自检问答工作表：思维模型自检问题。"""
    from openpyxl.styles import Font, Alignment
    ws_sc = wb.create_sheet("自检问答")

    questions = [
        ("01", "可以分解的总量指标是什么？可以按哪些维度分解？",
         "识别数据集中的连续总量指标（如销售额、订单量、成本）和可分解维度（如时间、品类、区域）"),
        ("02", "分解后，子群之间的行为差异有多大？倍数是多少？",
         "比较不同子群的关键指标差异，计算比率/倍数，评估差异的实际业务意义"),
        ("03", "哪些关键概念没有直接记录？可以用什么代理变量推断？",
         "如果核心概念（如用户价值、运营效率）没有直接字段，寻找可观测的替代指标"),
        ("04", "观察到的差异是用户偏好还是约束限制？",
         "通过数据判断差异来源：若是约束（不能选），建议调整供给；若是偏好（不愿选），建议调整营销"),
        ("05", "哪个细分领域既是高份额又存在显著差异（杠杆点）？",
         "寻找规模大且改进空间明显的子群，作为优先发力的业务切入点"),
    ]

    # 标题
    ws_sc.cell(row=1, column=1, value="思维模型自检问答").font = Font(bold=True, size=14)
    ws_sc.merge_cells("A1:D1")

    # 表头
    for c, h in enumerate(["编号", "自检问题", "参考答案", "本次分析应用"], 1):
        cell = ws_sc.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # 数据行
    for i, (num, question, answer) in enumerate(questions, 3):
        ws_sc.cell(row=i, column=1, value=num).border = thin_border
        cell_q = ws_sc.cell(row=i, column=2, value=question)
        cell_q.border = thin_border
        cell_q.alignment = Alignment(wrap_text=True)
        cell_a = ws_sc.cell(row=i, column=3, value=answer)
        cell_a.border = thin_border
        cell_a.alignment = Alignment(wrap_text=True)

        # 从 analysis_text 提取相关应用
        app = ""
        if analysis_text:
            kw = question[:10]
            app = _extract_section(analysis_text, [kw], 300)
        cell_app = ws_sc.cell(row=i, column=4, value=app if app else "可在分析中补充")
        cell_app.border = thin_border
        cell_app.alignment = Alignment(wrap_text=True)

    ws_sc.column_dimensions["A"].width = 8
    ws_sc.column_dimensions["B"].width = 40
    ws_sc.column_dimensions["C"].width = 45
    ws_sc.column_dimensions["D"].width = 45


def generate_excel(path: str, repo_root: str, output_dir: str = "", analysis_text: str = "", mode: str = "") -> str:
    """生成数据分析 Excel 交付物，包含数据表、统计表、看板和分析报告。"""
    headers, rows = _load_csv_data(path, repo_root)
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════
    # Sheet 1: 原始数据
    # ══════════════════════════════════════════════
    ws_data = wb.active
    ws_data.title = "数据概览"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for c, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            cell = ws_data.cell(row=r, column=c, value=row.get(h, ""))
            cell.border = thin_border

    for c in range(1, len(headers) + 1):
        ws_data.column_dimensions[get_column_letter(c)].width = max(12, len(headers[c - 1]) + 4)

    # ══════════════════════════════════════════════
    # Sheet 2: 整体对比（关键指标均值对比）
    # ══════════════════════════════════════════════
    ws_stats = wb.create_sheet("整体对比")
    numeric_cols = _infer_numeric_cols(headers, rows)

    stats_rows = [["指标", "均值", "中位数", "标准差", "最小值", "最大值", "计数"]]
    for col in numeric_cols:
        vals = []
        for r in rows:
            v = r.get(col, "").strip()
            if v:
                try:
                    vals.append(float(v))
                except ValueError:
                    pass
        if len(vals) < 2:
            continue
        vals.sort()
        n = len(vals)
        mean_v = statistics.mean(vals)
        median_v = statistics.median(vals)
        stdev_v = statistics.stdev(vals) if n > 1 else 0.0
        min_v = vals[0]
        max_v = vals[-1]
        stats_rows.append([col, round(mean_v, 2), round(median_v, 2), round(stdev_v, 2), min_v, max_v, n])

    for r, row_data in enumerate(stats_rows, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws_stats.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border

    # 数值列图表
    if len(numeric_cols) >= 1:
        chart = BarChart()
        chart.type = "col"
        chart.title = "数值列对比（均值）"
        chart.y_axis.title = "均值"
        data_ref = Reference(ws_stats, min_col=2, min_row=1, max_row=len(stats_rows), max_col=2)
        cats_ref = Reference(ws_stats, min_col=1, min_row=2, max_row=len(stats_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 20
        chart.height = 12
        ws_stats.add_chart(chart, f"J1")

    for c in range(1, 8):
        ws_stats.column_dimensions[get_column_letter(c)].width = 16

    # ══════════════════════════════════════════════
    # Sheet 3: 维度分析（自动检测类别列，分组聚合）
    # ══════════════════════════════════════════════
    _add_dimension_sheet(wb, headers, rows, numeric_cols, header_font, header_fill, thin_border)

    # ══════════════════════════════════════════════
    # Sheet 4: 交叉分析（2D 透视表）
    # ══════════════════════════════════════════════
    _add_cross_analysis_sheet(wb, headers, rows, numeric_cols, header_font, header_fill, thin_border)

    # ══════════════════════════════════════════════
    # Sheet 5: 思维模型（teaching content）
    # ══════════════════════════════════════════════
    _add_thinking_models_sheet(wb, analysis_text, header_font, header_fill, thin_border)

    # ══════════════════════════════════════════════
    # Sheet 6: 自检问答（Self-check Q&A）
    # ══════════════════════════════════════════════
    _add_self_check_sheet(wb, analysis_text, header_font, header_fill, thin_border)

    # ══════════════════════════════════════════════
    # Sheet 7: 分析看板（含执行摘要、核心发现、行动建议、KPI、关键指标）
    # ══════════════════════════════════════════════
    ws_dash = wb.create_sheet("分析看板")

    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell = ws_dash.cell(row=1, column=1, value="数据分析看板")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center")
    ws_dash.merge_cells("A1:F1")
    for c in range(1, 7):
        ws_dash.cell(row=1, column=c).fill = title_fill

    section_font = Font(bold=True, size=12, color="4472C4")
    current_row = 3

    # ── 执行摘要 ──
    structured_data = _extract_structured_summary(analysis_text) if analysis_text else {}

    if structured_data.get("summary"):
        ws_dash.cell(row=current_row, column=1, value="执行摘要").font = section_font
        current_row += 1
        cell = ws_dash.cell(row=current_row, column=1, value=structured_data["summary"])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_dash.merge_cells(f"A{current_row}:F{current_row}")
        ws_dash.row_dimensions[current_row].height = max(60, 15 * structured_data["summary"].count("\n") + 30)
        current_row += 2

    # ── 核心发现 ──
    findings = structured_data.get("findings", [])
    if findings:
        ws_dash.cell(row=current_row, column=1, value="Top 3 关键发现").font = section_font
        current_row += 1
        for i, finding in enumerate(findings[:3], 1):
            cell = ws_dash.cell(row=current_row, column=1, value=finding)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws_dash.merge_cells(f"A{current_row}:F{current_row}")
            current_row += 1
        current_row += 1

    # ── 行动建议 ──
    recommendations = structured_data.get("recommendations", "")
    if recommendations:
        ws_dash.cell(row=current_row, column=1, value="行动建议").font = section_font
        current_row += 1
        cell = ws_dash.cell(row=current_row, column=1, value=recommendations)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_dash.merge_cells(f"A{current_row}:F{current_row}")
        ws_dash.row_dimensions[current_row].height = max(60, 15 * recommendations.count("\n") + 30)
        current_row += 2

    # ── KPI 卡片 ──
    kpi_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    kpi_font = Font(bold=True, size=11)

    ws_dash.cell(row=current_row, column=1, value="数据集").font = kpi_font
    ws_dash.cell(row=current_row, column=1).fill = kpi_fill
    ws_dash.cell(row=current_row, column=2, value=Path(path).name).fill = kpi_fill
    ws_dash.merge_cells(f"B{current_row}:F{current_row}")
    current_row += 1

    ws_dash.cell(row=current_row, column=1, value="总行数").font = kpi_font
    ws_dash.cell(row=current_row, column=1).fill = kpi_fill
    ws_dash.cell(row=current_row, column=2, value=len(rows)).fill = kpi_fill
    current_row += 1

    ws_dash.cell(row=current_row, column=1, value="总列数").font = kpi_font
    ws_dash.cell(row=current_row, column=1).fill = kpi_fill
    ws_dash.cell(row=current_row, column=2, value=len(headers)).fill = kpi_fill
    current_row += 2

    # ── 关键指标 ──
    if numeric_cols:
        ws_dash.cell(row=current_row, column=1, value="关键指标").font = Font(bold=True, size=12)
        current_row += 1
        for i, col in enumerate(numeric_cols[:5]):
            vals = []
            for r in rows:
                v = r.get(col, "").strip()
                if v:
                    try:
                        vals.append(float(v))
                    except ValueError:
                        pass
            if len(vals) < 2:
                continue
            ws_dash.cell(row=current_row, column=1, value=col)
            ws_dash.cell(row=current_row, column=2, value=f"均值: {statistics.mean(vals):.2f}")
            ws_dash.cell(row=current_row, column=3, value=f"合计: {sum(vals):.2f}")
            current_row += 1

    for c in range(1, 7):
        ws_dash.column_dimensions[get_column_letter(c)].width = 80

    # ══════════════════════════════════════════════
    # Sheet 8: 分析报告（agent 分析结论）
    # ══════════════════════════════════════════════
    if analysis_text:
        ws_report = wb.create_sheet("分析报告")
        # 把分析文本按行写入，每行一个单元格
        report_lines = analysis_text.split("\n")
        for r, line in enumerate(report_lines, 1):
            cell = ws_report.cell(row=r, column=1, value=line)
            if r == 1:
                cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(wrap_text=True)
        ws_report.column_dimensions["A"].width = 120

    # 保存
    output_name = Path(path).stem
    output_path = Path(output_dir) / f"{output_name}_analysis.xlsx" if output_dir else Path(path).parent / f"{output_name}_analysis.xlsx"
    wb.save(str(output_path))
    return f"✅ Excel 已生成: {output_path}"


def _analysis_to_html(text: str) -> str:
    """将 agent 分析报告文本转为安全的 HTML 区块。所有文本内容经过转义，保留中文。"""
    lines = text.strip().split("\n")
    html_parts = ['<h2>分析结论</h2>', '<div class="analysis-content">']
    in_list = False

    def _close_list():
        nonlocal in_list
        if in_list:
            html_parts.append('</ul>')
            in_list = False

    for line in lines:
        stripped = line.strip()
        if not stripped:
            _close_list()
            html_parts.append('<br>')
        elif stripped.startswith("## ") or stripped.startswith("# "):
            _close_list()
            level = 2 if stripped.startswith("##") else 3
            title = html.escape(stripped.lstrip("#").strip())
            html_parts.append(f'<h{level}>{title}</h{level}>')
        elif re.match(r'^\|\s*[-:| ]+\s*\|', stripped):
            continue  # skip markdown table separator rows
        elif stripped.startswith("|"):
            _close_list()
            cells = [html.escape(c.strip()) for c in stripped.strip("|").split("|")]
            html_parts.append("<tr>" + "".join(f"<td>{c}</td>" for c in cells) + "</tr>")
        elif stripped.startswith("- "):
            if not in_list:
                html_parts.append('<ul>')
                in_list = True
            cleaned = stripped[2:].replace('**', '')
            html_parts.append(f'<li>{html.escape(cleaned)}</li>')
        elif re.match(r'^\d+[\.\)]\s', stripped):
            if not in_list:
                html_parts.append('<ol>')
                in_list = True
            cleaned = stripped.replace('**', '')
            html_parts.append(f'<li>{html.escape(cleaned)}</li>')
        elif "**" in stripped:
            _close_list()
            # Convert **bold** to <strong> while escaping
            processed = html.escape(stripped)
            processed = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', processed)
            html_parts.append(f'<p>{processed}</p>')
        else:
            _close_list()
            html_parts.append(f'<p>{html.escape(stripped)}</p>')

    _close_list()
    html_parts.append('</div>')
    return "\n".join(html_parts)


def generate_html(path: str, repo_root: str, output_dir: str = "", analysis_text: str = "", mode: str = "") -> str:
    """生成数据分析 HTML 报告，包含统计数据和 agent 分析结论。"""
    headers, rows = _load_csv_data(path, repo_root)
    numeric_cols = _infer_numeric_cols(headers, rows)

    # 统计计算
    stats_html = ""
    for col in numeric_cols:
        vals = []
        for r in rows:
            v = r.get(col, "").strip()
            if v:
                try:
                    vals.append(float(v))
                except ValueError:
                    pass
        if len(vals) < 2:
            continue
        n = len(vals)
        mean_v = statistics.mean(vals)
        stats_html += f"""
        <tr>
            <td>{col}</td>
            <td>{mean_v:.2f}</td>
            <td>{statistics.median(vals):.2f}</td>
            <td>{statistics.stdev(vals):.2f}</td>
            <td>{min(vals):.2f}</td>
            <td>{max(vals):.2f}</td>
            <td>{n}</td>
        </tr>"""

    # 前 5 行数据
    data_rows_html = ""
    for i, row in enumerate(rows[:10]):
        data_rows_html += "<tr>" + "".join(f"<td>{row.get(h, '')}</td>" for h in headers) + "</tr>"

    # 结构化摘要区块（描述性统计之后、完整报告之前）
    structured_html = ""
    if analysis_text:
        _sd = _extract_structured_summary(analysis_text)
        _parts = []
        if _sd.get("summary"):
            _parts.append('<h2>执行摘要</h2>')
            _parts.append(f'<p>{__import__("html").escape(_sd["summary"])}</p>')
        if _sd.get("findings"):
            _parts.append('<h2>核心发现</h2>')
            _parts.append('<ol>')
            for f in _sd["findings"]:
                _parts.append(f'<li>{__import__("html").escape(f)}</li>')
            _parts.append('</ol>')
        if _sd.get("recommendations"):
            _parts.append('<h2>行动建议</h2>')
            for line in _sd["recommendations"].split("\n"):
                line = line.strip()
                if line:
                    _parts.append(f'<p>{__import__("html").escape(line)}</p>')
        structured_html = "\n".join(_parts)

    dataset_name = Path(path).name
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>数据分析报告 - {dataset_name}</title>
<style>
body {{ font-family: -apple-system, "Segoe UI", Helvetica, Arial, sans-serif; max-width: 960px; margin: 0 auto; padding: 20px; color: #333; }}
h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
h2 {{ color: #2980b9; margin-top: 30px; }}
table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
th, td {{ border: 1px solid #ddd; padding: 8px 12px; text-align: left; }}
th {{ background-color: #3498db; color: white; font-weight: 600; }}
tr:nth-child(even) {{ background-color: #f8f9fa; }}
.kpi {{ display: inline-block; background: #e8f4f8; border-radius: 8px; padding: 15px 25px; margin: 10px; text-align: center; }}
.kpi-value {{ font-size: 28px; font-weight: bold; color: #2980b9; }}
.kpi-label {{ font-size: 13px; color: #666; }}
.footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #eee; color: #999; font-size: 13px; }}
</style>
</head>
<body>
<h1>数据分析报告</h1>
<p>数据集: <strong>{dataset_name}</strong></p>

<div class="kpi"><div class="kpi-value">{len(rows)}</div><div class="kpi-label">数据行数</div></div>
<div class="kpi"><div class="kpi-value">{len(headers)}</div><div class="kpi-label">字段数量</div></div>
<div class="kpi"><div class="kpi-value">{len(numeric_cols)}</div><div class="kpi-label">数值列</div></div>

<h2>字段信息</h2>
<table><tr>{"".join(f'<th>{h}</th>' for h in headers)}</tr>{data_rows_html}</table>

<h2>描述性统计</h2>
<table>
<tr><th>指标</th><th>均值</th><th>中位数</th><th>标准差</th><th>最小值</th><th>最大值</th><th>计数</th></tr>
{stats_html}
</table>

{structured_html}

{_analysis_to_html(analysis_text) if analysis_text else ""}
<div class="footer">
<p>生成工具: DataHelp | 生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
</div>
</body>
</html>"""

    output_name = Path(path).stem
    output_path = Path(output_dir) / f"{output_name}_report.html" if output_dir else Path(path).parent / f"{output_name}_report.html"
    output_path.write_text(html, encoding="utf-8")
    return f"✅ HTML 已生成: {output_path}"


def _find_cjk_font() -> str:
    """解析 CJK 字体路径，供 generate_pdf 注册使用。

    优先级：
    1. DATAHELP_CJK_FONT_PATH 环境变量
    2. 预定义候选路径列表

    Returns:
        字体文件绝对路径

    Raises:
        RuntimeError: 找不到任何可用的 CJK 字体
    """
    env_path = os.environ.get("DATAHELP_CJK_FONT_PATH")
    if env_path:
        resolved = os.path.abspath(env_path)
        if os.path.exists(resolved):
            return resolved

    candidates = [
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Hiragino Sans GB.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttf",
    ]
    for fp in candidates:
        if os.path.exists(fp):
            return fp

    raise RuntimeError(
        "未找到 CJK 字体。请设置环境变量 DATAHELP_CJK_FONT_PATH 指向中文字体文件，"
        "或安装 Noto Sans CJK 字体。"
    )


def generate_pdf(path: str, repo_root: str, output_dir: str = "", analysis_text: str = "", mode: str = "") -> str:
    """生成数据分析 PDF 报告，使用 CJK 字体渲染中文内容。"""
    from fpdf import FPDF

    cjk_font_path = _find_cjk_font()
    cjk_family = "DataHelpCJK"

    headers, rows = _load_csv_data(path, repo_root)
    numeric_cols = _infer_numeric_cols(headers, rows)
    dataset_name = Path(path).name
    total_rows = len(rows)
    total_cols = len(headers)

    # 解析报告章节
    analysis_text = analysis_text.replace("⚠", "注意：").replace("️", "")
    report_sections = _parse_report_sections(analysis_text)
    structured_summary = _extract_structured_summary(analysis_text) if analysis_text else {}

    # 同步清理 structured_summary 中可能含 ⚠ 和 markdown blockquote 前缀的文本
    def _strip_blockquote(text: str) -> str:
        """清除每行的 '> ' 块引用前缀。"""
        return "\n".join(
            line[2:] if line.startswith("> ") else line
            for line in text.split("\n")
        )

    if structured_summary.get("summary"):
        text = structured_summary["summary"].replace("⚠", "注意：")
        structured_summary["summary"] = _strip_blockquote(text)
    for i, f in enumerate(structured_summary.get("findings", [])):
        text = f.replace("⚠", "注意：")
        structured_summary["findings"][i] = _strip_blockquote(text)
    if structured_summary.get("recommendations"):
        text = structured_summary["recommendations"].replace("⚠", "注意：")
        structured_summary["recommendations"] = _strip_blockquote(text)

    pdf = FPDF()
    pdf.add_page()

    # 注册 CJK 字体（同一字体文件同时注册 regular 和 bold 样式）
    pdf.add_font(cjk_family, "", cjk_font_path)
    # fpdf2 会对同一字体文件做合成粗体处理
    pdf.add_font(cjk_family, "B", cjk_font_path)

    # ── 中文封面 ──
    pdf.set_font(cjk_family, "B", 24)
    pdf.cell(0, 15, "数据分析报告", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)
    pdf.set_font(cjk_family, "", 13)
    pdf.cell(0, 10, f"数据集: {dataset_name}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.cell(0, 10, f"行数: {total_rows}  |  列数: {total_cols}", new_x="LMARGIN", new_y="NEXT", align="C")
    if mode:
        pdf.cell(0, 10, f"模式: {mode}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(10)

    # ── 执行摘要（只写 helper 的摘要） ──
    pdf.set_font(cjk_family, "B", 16)
    pdf.cell(0, 12, "执行摘要", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(cjk_family, "", 10)
    if structured_summary.get("summary"):
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(pdf.epw, 6, structured_summary["summary"], new_x="LMARGIN")
    else:
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(pdf.epw, 6, f"本报告分析了 {dataset_name}，共 {total_rows} 行数据，{total_cols} 个字段。", new_x="LMARGIN")
    pdf.ln(5)

    # ── 关键指标表 ──
    pdf.set_font(cjk_family, "B", 14)
    pdf.cell(0, 12, "关键指标", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(cjk_family, "", 10)
    col_w = pdf.w / (min(len(numeric_cols) + 1, 7))
    pdf.cell(col_w, 8, "列名", 1)
    pdf.cell(col_w, 8, "均值", 1)
    pdf.cell(col_w, 8, "中位数", 1)
    pdf.cell(col_w, 8, "标准差", 1)
    pdf.cell(col_w, 8, "最小值", 1)
    pdf.cell(col_w, 8, "最大值", 1)
    pdf.ln()

    pdf.set_font(cjk_family, "", 9)
    for col in numeric_cols[:8]:
        vals = []
        for r in rows:
            v = r.get(col, "").strip()
            if v:
                try:
                    vals.append(float(v))
                except ValueError:
                    pass
        if len(vals) < 2:
            continue
        pdf.cell(col_w, 7, col[:12], 1)
        pdf.cell(col_w, 7, f"{statistics.mean(vals):.1f}", 1)
        pdf.cell(col_w, 7, f"{statistics.median(vals):.1f}", 1)
        pdf.cell(col_w, 7, f"{statistics.stdev(vals):.1f}", 1)
        pdf.cell(col_w, 7, f"{min(vals):.1f}", 1)
        pdf.cell(col_w, 7, f"{max(vals):.1f}", 1)
        pdf.ln()
        if pdf.get_y() > 260:
            pdf.add_page()

    # ── 核心发现（使用 helper 结果） ──
    findings = structured_summary.get("findings", [])
    if findings:
        pdf.add_page()
        pdf.set_font(cjk_family, "B", 16)
        pdf.cell(0, 12, "核心发现", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(cjk_family, "", 10)
        for finding in findings:
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(pdf.epw, 6, finding, new_x="LMARGIN")
            pdf.ln(3)
            if pdf.get_y() > 260:
                pdf.add_page()

    # ── 行动建议（使用 helper 结果） ──
    recommendations = structured_summary.get("recommendations", "")
    if recommendations:
        pdf.add_page()
        pdf.set_font(cjk_family, "B", 16)
        pdf.cell(0, 12, "行动建议", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)
        pdf.set_font(cjk_family, "", 10)
        for line in recommendations.split("\n"):
            stripped = line.strip()
            if not stripped:
                pdf.ln(3)
            else:
                pdf.set_x(pdf.l_margin)
                pdf.multi_cell(pdf.epw, 6, stripped, new_x="LMARGIN")
            if pdf.get_y() > 260:
                pdf.add_page()

    # ── 思维模型应用（仅 audit_report 模式） ──
    if analysis_text and mode == "audit_report":
        pdf.add_page()
        pdf.set_font(cjk_family, "",16)
        pdf.cell(0, 12, "思维模型应用", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)
        thinking_models = [
            ("1. 分解", "把总量指标拆解为有意义的组成部分（按时间、品类、区域等）"),
            ("2. 分层差异", "不盲目相信平均值，比较子群间的差异（倍数/比率）"),
            ("3. 代理推断", "当关键概念没有直接记录时，用可观测的代理信号推断"),
            ("4. 约束 vs 偏好", "差异来自用户不能选择（约束）还是不愿选择（偏好）"),
            ("5. 杠杆点", "聚焦高份额 + 高改进空间的细分领域"),
        ]
        for name, desc in thinking_models:
            pdf.set_font(cjk_family, "B", 11)
            pdf.cell(0, 8, name, new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(cjk_family, "", 10)
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(pdf.epw, 6, desc, new_x="LMARGIN")
            kw = name.split(".", 1)[1].strip() if "." in name else name
            evidence = _extract_section(analysis_text, [kw, desc[:12]], 200)
            if evidence:
                pdf.set_font(cjk_family, "", 9)
                pdf.set_x(pdf.l_margin)
                pdf.multi_cell(pdf.epw, 5, f"数据证据: {evidence[:200]}", new_x="LMARGIN")
            pdf.ln(2)

    # ── 数据概览页 ──
    if len(rows) > 0:
        pdf.add_page()
        pdf.set_font(cjk_family, "",14)
        pdf.cell(0, 10, "数据预览", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(cjk_family, "", 7)
        col_w2 = pdf.w / (min(len(headers) + 1, 8))
        for h in headers[:7]:
            pdf.cell(col_w2, 7, h[:10], 1)
        pdf.ln()
        pdf.set_font(cjk_family, "", 7)
        for i, row in enumerate(rows[:25]):
            if pdf.get_y() > 260:
                pdf.add_page()
            for h in headers[:7]:
                pdf.cell(col_w2, 6, str(row.get(h, ""))[:10], 1)
            pdf.ln()

    # ── 完整分析报告 ──
    if analysis_text:
        pdf.add_page()
        pdf.set_font(cjk_family, "",16)
        pdf.cell(0, 15, "完整分析报告", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(cjk_family, "", 10)
        for line in analysis_text.split("\n"):
            stripped = line.strip()
            if not stripped:
                pdf.ln(3)
            elif stripped.startswith("##") or stripped.startswith("#"):
                title = stripped.lstrip("#").strip()
                pdf.set_font(cjk_family, "B", 12)
                pdf.set_x(pdf.l_margin)
                pdf.multi_cell(pdf.epw, 8, title, new_x="LMARGIN")
                pdf.set_font(cjk_family, "", 10)
            elif stripped.startswith("|"):
                continue
            elif stripped.startswith("> "):
                cleaned = _clean_markdown(stripped[2:].strip())
                if cleaned:
                    pdf.set_x(pdf.l_margin)
                    pdf.multi_cell(pdf.epw, 5, cleaned, new_x="LMARGIN")
            else:
                cleaned = _clean_markdown(stripped)
                if cleaned:
                    pdf.set_x(pdf.l_margin)
                    pdf.multi_cell(pdf.epw, 5, cleaned, new_x="LMARGIN")
            if pdf.get_y() > 260:
                pdf.add_page()

    output_name = Path(path).stem
    output_path = Path(output_dir) / f"{output_name}_report.pdf" if output_dir else Path(path).parent / f"{output_name}_report.pdf"
    pdf.output(str(output_path))
    return f"✅ PDF 已生成: {output_path}"
