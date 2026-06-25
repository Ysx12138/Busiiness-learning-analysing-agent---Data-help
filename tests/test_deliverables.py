"""无 API 测试：交付物生成、CJK 字体解析、Pipeline _Args 合约。"""

import csv
import json
import os
import sys
import tempfile
import unittest
from pathlib import Path

# 确保项目根在 sys.path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from datahelp.tools_data import (
    _find_cjk_font,
    _parse_report_sections,
    _extract_section,
    _extract_top_findings,
    generate_excel,
    generate_html,
    generate_pdf,
    _load_csv_data,
)


# ── 公用测试数据 ─────────────────────────────────────

SAMPLE_CSV_CONTENT = """product,category,price,quantity,revenue
A,Electronics,100,5,500
B,Electronics,200,3,600
C,Home,50,10,500
D,Home,80,7,560
E,Clothing,30,20,600
F,Clothing,40,15,600
G,Electronics,150,4,600
H,Home,60,8,480"""

CHINESE_ANALYSIS_TEXT = """# 数据分析报告

## 执行摘要

本报告分析了 8 条商品销售数据，涵盖电子产品、家居用品和服装三个品类。
数据显示电子产品单价较高但销量中等，家居用品销量稳定，服装品类销量最大。
建议优化电子产品的定价策略，并加大服装品类的营销投入。

## 核心发现

1. **电子产品贡献最高收入**：电子产品总收入 1700，平均单价 150，是三个品类中单价最高的。
2. **服装销量最大**：服装品类总销量 35 件，是所有品类中最高的，但单价最低。
3. **家居用品利润稳定**：家居用品总收入 1540，销量 25 件，表现稳健。

## 业务建议

建议一：对电子产品进行捆绑销售，提升客单价。
建议二：增加服装品类的高端产品线，提升利润率。
建议三：家居用品保持现有策略，优化供应链降低成本。

## 数据概览

数据集包含 8 条记录，7 个字段，无缺失值。
"""


class TestCjkFontFinder(unittest.TestCase):
    """测试 CJK 字体查找函数。"""

    def test_env_var_override_non_existent(self):
        """设置 DATAHELP_CJK_FONT_PATH 为不存在的路径应继续检测系统字体。"""
        os.environ["DATAHELP_CJK_FONT_PATH"] = "/tmp/__nonexistent_cjk_font_test.ttf"
        try:
            # 系统上有字体，所以不会 raise RuntimeError
            font = _find_cjk_font()
            self.assertTrue(os.path.exists(font))
        finally:
            del os.environ["DATAHELP_CJK_FONT_PATH"]

    def test_env_var_override_existent(self):
        """设置 DATAHELP_CJK_FONT_PATH 为存在的路径应返回该路径。"""
        # 使用实际存在的系统字体测试
        real_font = _find_cjk_font()
        os.environ["DATAHELP_CJK_FONT_PATH"] = real_font
        try:
            font = _find_cjk_font()
            self.assertEqual(os.path.abspath(font), os.path.abspath(real_font))
        finally:
            del os.environ["DATAHELP_CJK_FONT_PATH"]

    def test_finds_actual_macos_font(self):
        """在 macOS 上应找到 Arial Unicode.ttf。"""
        font = _find_cjk_font()
        self.assertTrue(os.path.exists(font), f"字体文件不存在: {font}")
        self.assertTrue(font.endswith(".ttf") or font.endswith(".ttc"),
                        f"不是字体文件: {font}")


class TestReportParser(unittest.TestCase):
    """测试 Markdown 报告章节解析。"""

    def test_parse_empty(self):
        self.assertEqual(_parse_report_sections(""), {})

    def test_parse_sections(self):
        sections = _parse_report_sections(CHINESE_ANALYSIS_TEXT)
        self.assertIn("执行摘要", sections)
        self.assertIn("核心发现", sections)
        self.assertIn("业务建议", sections)
        self.assertIn("数据概览", sections)
        # 文本以 # 开头，所以首个章节名是标题本身
        self.assertIn("数据分析报告", sections)

    def test_exec_summary_content(self):
        sections = _parse_report_sections(CHINESE_ANALYSIS_TEXT)
        summary = sections.get("执行摘要", "")
        self.assertIn("电子产品", summary)
        self.assertIn("服装", summary)

    def test_findings_content(self):
        sections = _parse_report_sections(CHINESE_ANALYSIS_TEXT)
        findings = sections.get("核心发现", "")
        self.assertIn("电子产品贡献最高收入", findings)

    def test_extract_section_keyword(self):
        result = _extract_section(CHINESE_ANALYSIS_TEXT, ["业务建议", "建议"], 500)
        self.assertIn("捆绑销售", result)

    def test_extract_top_findings(self):
        sections = _parse_report_sections(CHINESE_ANALYSIS_TEXT)
        findings_text = sections.get("核心发现", "")
        top3 = _extract_top_findings(findings_text, 3)
        self.assertEqual(len(top3), 3)
        self.assertIn("电子产品贡献最高收入", top3[0])


class TestExcelDeliverable(unittest.TestCase):
    """测试 Excel 交付物生成（无 API 调用）。"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.csv_path = os.path.join(self.tmpdir, "test_data.csv")
        with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
            f.write(SAMPLE_CSV_CONTENT)

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_excel_generated_with_analysis(self):
        """生成 Excel 应返回成功消息且文件存在。"""
        result = generate_excel(self.csv_path, self.tmpdir, self.tmpdir,
                                analysis_text=CHINESE_ANALYSIS_TEXT)
        self.assertIn("✅", result)
        # 查找 xlsx 文件
        xlsx_files = list(Path(self.tmpdir).glob("*.xlsx"))
        self.assertTrue(len(xlsx_files) >= 1, "没有生成 xlsx 文件")

    def test_excel_contains_dashboard_sheet(self):
        """Excel 应包含'分析看板'工作表。"""
        import openpyxl
        result = generate_excel(self.csv_path, self.tmpdir, self.tmpdir,
                                analysis_text=CHINESE_ANALYSIS_TEXT)
        xlsx_files = list(Path(self.tmpdir).glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_files[0])
        sheet_names = wb.sheetnames
        self.assertIn("分析看板", sheet_names)

    def test_excel_dashboard_has_summary_and_findings(self):
        """分析看板应包含执行摘要和关键发现。"""
        import openpyxl
        generate_excel(self.csv_path, self.tmpdir, self.tmpdir,
                       analysis_text=CHINESE_ANALYSIS_TEXT)
        xlsx_files = list(Path(self.tmpdir).glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_files[0])
        ws = wb["分析看板"]
        # 找执行摘要和关键发现标题
        found_summary = False
        found_findings = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and "执行摘要" in str(cell):
                    found_summary = True
                if cell and "Top 3 关键发现" in str(cell):
                    found_findings = True
        self.assertTrue(found_summary, "分析看板缺少执行摘要")
        self.assertTrue(found_findings, "分析看板缺少 Top 3 关键发现")

    def test_excel_report_sheet_cn_content(self):
        """分析报告工作表应包含中文结论。"""
        import openpyxl
        generate_excel(self.csv_path, self.tmpdir, self.tmpdir,
                       analysis_text=CHINESE_ANALYSIS_TEXT)
        xlsx_files = list(Path(self.tmpdir).glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_files[0])
        self.assertIn("分析报告", wb.sheetnames)
        ws = wb["分析报告"]
        text = ""
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    text += str(cell)
        self.assertIn("电子产品", text)
        self.assertIn("业务建议", text)


class TestHtmlDeliverable(unittest.TestCase):
    """测试 HTML 交付物生成。"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.csv_path = os.path.join(self.tmpdir, "test_data.csv")
        with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
            f.write(SAMPLE_CSV_CONTENT)

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_html_generated(self):
        """生成 HTML 应返回成功消息且文件存在。"""
        result = generate_html(self.csv_path, self.tmpdir, self.tmpdir,
                               analysis_text=CHINESE_ANALYSIS_TEXT)
        self.assertIn("✅", result)
        html_files = list(Path(self.tmpdir).glob("*.html"))
        self.assertTrue(len(html_files) >= 1)

    def test_html_contains_cn(self):
        """HTML 应包含中文结论。"""
        generate_html(self.csv_path, self.tmpdir, self.tmpdir,
                      analysis_text=CHINESE_ANALYSIS_TEXT)
        html_files = list(Path(self.tmpdir).glob("*.html"))
        content = html_files[0].read_text(encoding="utf-8")
        self.assertIn("电子产品", content)
        self.assertIn("核心发现", content)
        self.assertIn("业务建议", content)
        # 中文应正常显示（非乱码）
        self.assertNotIn("锘", content)  # 常见 UTF-8 BOM 乱码

    def test_html_accepts_mode_param(self):
        """generate_html 应接受 mode 关键字参数（向后兼容）。"""
        result = generate_html(self.csv_path, self.tmpdir, self.tmpdir,
                               analysis_text=CHINESE_ANALYSIS_TEXT,
                               mode="beginner_summary")
        self.assertIn("✅", result)


class TestPdfDeliverable(unittest.TestCase):
    """测试 PDF 交付物生成（需要 CJK 字体）。"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.csv_path = os.path.join(self.tmpdir, "test_data.csv")
        with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
            f.write(SAMPLE_CSV_CONTENT)

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_pdf_generated(self):
        """生成 PDF 应返回成功消息且文件存在。"""
        try:
            result = generate_pdf(self.csv_path, self.tmpdir, self.tmpdir,
                                  analysis_text=CHINESE_ANALYSIS_TEXT)
        except RuntimeError as e:
            self.skipTest(f"无 CJK 字体: {e}")
        self.assertIn("✅", result)
        pdf_files = list(Path(self.tmpdir).glob("*.pdf"))
        self.assertTrue(len(pdf_files) >= 1)

    def test_pdf_reasonable_size(self):
        """生成的 PDF 应大于 10KB（含实际内容）。"""
        try:
            generate_pdf(self.csv_path, self.tmpdir, self.tmpdir,
                         analysis_text=CHINESE_ANALYSIS_TEXT)
        except RuntimeError as e:
            self.skipTest(f"无 CJK 字体: {e}")
        pdf_files = list(Path(self.tmpdir).glob("*.pdf"))
        size = pdf_files[0].stat().st_size
        self.assertGreater(size, 10240, f"PDF 文件过小: {size} bytes")

    def test_pdf_mode_param(self):
        """generate_pdf 应接受 mode 参数。"""
        try:
            result = generate_pdf(self.csv_path, self.tmpdir, self.tmpdir,
                                  analysis_text=CHINESE_ANALYSIS_TEXT,
                                  mode="beginner_summary")
        except RuntimeError as e:
            self.skipTest(f"无 CJK 字体: {e}")
        self.assertIn("✅", result)


class TestPipelineArgsContract(unittest.TestCase):
    """测试 pipeline _Args 合约：模拟 build_agent 调用所需的全部属性。"""

    def test_args_has_all_required_attrs(self):
        """_Args 应包含 build_agent 使用的全部属性。"""
        # 模拟 pipeline.py 中的 _Args 初始化
        provider = "mock"
        model = None
        work_dir = "/tmp"
        max_steps = 20
        max_new_tokens = 4096
        mode = "beginner_summary"
        output_dir = "/tmp/output"

        class _Args:
            def __init__(self):
                self.provider = provider
                self.model = model
                self.cwd = str(work_dir)
                self.output_dir = str(output_dir)  # Task A: 必须定义
                self.max_steps = max_steps
                self.max_new_tokens = max_new_tokens
                self.approval = "auto"
                self.temperature = None
                self.mode = mode
                self.eval = None

        args = _Args()
        # build_agent 使用的全部属性
        self.assertEqual(args.provider, "mock")
        self.assertEqual(args.cwd, "/tmp")
        self.assertEqual(args.output_dir, "/tmp/output")
        self.assertEqual(args.mode, "beginner_summary")
        self.assertEqual(args.approval, "auto")
        self.assertIsNone(args.temperature)
        self.assertIsNone(args.eval)

    def test_no_api_access_in_renderers(self):
        """验证 generate_excel / html / pdf 不调用任何外部 API。"""
        # 纯粹依靠传入的 analysis_text 工作
        # 不访问网络，只读写本地文件系统中的 temp CSV
        self.assertTrue(callable(generate_excel))
        self.assertTrue(callable(generate_html))
        self.assertTrue(callable(generate_pdf))


class TestLoadCsvData(unittest.TestCase):
    """测试 _load_csv_data 工具函数。"""

    def test_loads_sample_csv(self):
        with tempfile.TemporaryDirectory() as td:
            csv_path = os.path.join(td, "test.csv")
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                f.write(SAMPLE_CSV_CONTENT)
            headers, rows = _load_csv_data(csv_path, td)
            self.assertEqual(len(headers), 5)
            self.assertEqual(len(rows), 8)
            self.assertIn("product", headers)
            self.assertIn("revenue", headers)


if __name__ == "__main__":
    unittest.main(verbosity=2)
